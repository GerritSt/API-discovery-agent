"""
Excel Export Module
Exports AI-discovered API information to Excel spreadsheets.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict
import logging
from datetime import datetime
import os

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports AI-discovered API information to Excel spreadsheets."""
    
    def __init__(self):
        """Initialize the Excel exporter."""
        self.workbook = None
        
    def create_spreadsheet(self, api_info: Dict, filename: str = None) -> str:
        """
        Create an Excel spreadsheet with API information discovered by AI.
        
        Args:
            api_info: Dictionary containing API information from AI discovery
            filename: Optional output filename (auto-generated if not provided)
            
        Returns:
            Path to the created Excel file
        """
        company_name = api_info.get('company_name', 'Unknown')
        logger.info(f"Creating Excel spreadsheet for {company_name}")
        
        # Generate filename
        if not filename:
            filename = f"{company_name.replace(' ', '_')}_API_Endpoints.xlsx"
        
        # Ensure file saves to data directory
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(current_dir)
        data_dir = os.path.join(project_root, 'data')
        os.makedirs(data_dir, exist_ok=True)
        
        if not os.path.isabs(filename):
            filename = os.path.join(data_dir, filename)
        
        # Create workbook
        self.workbook = Workbook()
        sheet = self.workbook.active
        sheet.title = "API Information"
        
        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=16)
        
        # Title
        sheet.merge_cells('A1:B1')
        title_cell = sheet['A1']
        title_cell.value = f"{company_name} API Information"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Metadata section
        row = 3
        sheet[f'A{row}'] = "Has API:"
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'] = "Yes" if api_info.get('has_api') else "No"
        
        row += 1
        sheet[f'A{row}'] = "API Type:"
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'] = api_info.get('api_type', 'N/A')
        
        row += 1
        sheet[f'A{row}'] = "Base URL:"
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'] = api_info.get('base_url', 'N/A')
        
        row += 1
        sheet[f'A{row}'] = "Generated:"
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # API Endpoints section
        row += 2
        sheet.merge_cells(f'A{row}:C{row}')
        header_cell = sheet[f'A{row}']
        header_cell.value = "API Endpoints"
        header_cell.font = Font(color="FFFFFF", bold=True, size=12)
        header_cell.fill = header_fill
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'B{row}'].fill = header_fill
        sheet[f'C{row}'].fill = header_fill
        
        row += 1
        sheet[f'A{row}'] = "Method"
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        sheet[f'B{row}'] = "Path"
        sheet[f'B{row}'].font = Font(bold=True)
        sheet[f'B{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        sheet[f'C{row}'] = "Description"
        sheet[f'C{row}'].font = Font(bold=True)
        sheet[f'C{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Add endpoints (sorted by path for better organization)
        endpoints = api_info.get('endpoints', [])
        # Sort endpoints by path, then by method
        sorted_endpoints = sorted(endpoints, key=lambda x: (x.get('path', ''), x.get('method', '')))
        
        for endpoint in sorted_endpoints:
            row += 1
            sheet[f'A{row}'] = endpoint.get('method', '')
            sheet[f'B{row}'] = endpoint.get('path', '')
            sheet[f'C{row}'] = endpoint.get('description', '')
        
        # Column widths
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 60
        
        # Save workbook
        self.workbook.save(filename)
        logger.info(f"Excel file saved: {filename}")
        
        return filename

