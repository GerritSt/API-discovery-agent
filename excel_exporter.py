"""
Excel Export Module
Handles exporting API endpoint data to Excel spreadsheets.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports API endpoint information to Excel spreadsheets."""
    
    def __init__(self):
        self.workbook = None
        
    def create_spreadsheet(self, company_name: str, doc_url: str, endpoints: List[Dict[str, str]], filename: str = None) -> str:
        """
        Create an Excel spreadsheet with API endpoint information.
        
        Args:
            company_name: Name of the company
            doc_url: URL of the API documentation
            endpoints: List of endpoint dictionaries
            filename: Output filename (optional)
            
        Returns:
            Path to the created Excel file
        """
        logger.info(f"Creating Excel spreadsheet for {company_name}")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{company_name.replace(' ', '_')}_API_Documentation_{timestamp}.xlsx"
        
        # Create workbook and select active sheet
        self.workbook = Workbook()
        sheet = self.workbook.active
        sheet.title = "API Endpoints"
        
        # Set up header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Add title section
        sheet.merge_cells('A1:E1')
        title_cell = sheet['A1']
        title_cell.value = f"{company_name} API Documentation"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add metadata
        sheet['A2'] = "Documentation URL:"
        sheet['B2'] = doc_url
        sheet['B2'].font = Font(color="0000FF", underline="single")
        
        sheet['A3'] = "Total Endpoints:"
        sheet['B3'] = len(endpoints)
        
        sheet['A4'] = "Generated:"
        sheet['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Add headers
        headers = ['Method', 'Path', 'Full Endpoint', 'Description', 'Notes']
        header_row = 6
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add endpoint data
        current_row = header_row + 1
        for endpoint in endpoints:
            sheet.cell(row=current_row, column=1, value=endpoint.get('method', ''))
            sheet.cell(row=current_row, column=2, value=endpoint.get('path', ''))
            sheet.cell(row=current_row, column=3, value=endpoint.get('full_endpoint', ''))
            sheet.cell(row=current_row, column=4, value=endpoint.get('description', ''))
            sheet.cell(row=current_row, column=5, value='')  # Empty notes column for user input
            
            # Apply alternating row colors
            if current_row % 2 == 0:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for col in range(1, 6):
                    sheet.cell(row=current_row, column=col).fill = fill
            
            current_row += 1
        
        # Adjust column widths
        column_widths = {
            'A': 12,  # Method
            'B': 40,  # Path
            'C': 50,  # Full Endpoint
            'D': 60,  # Description
            'E': 30   # Notes
        }
        
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width
        
        # Freeze the header row
        sheet.freeze_panes = 'A7'
        
        # Add filters
        sheet.auto_filter.ref = f"A{header_row}:E{current_row - 1}"
        
        # Save the workbook
        self.workbook.save(filename)
        logger.info(f"Excel file saved: {filename}")
        
        return filename
    
    def add_summary_sheet(self, sheet_name: str = "Summary"):
        """Add a summary sheet to the workbook."""
        if not self.workbook:
            logger.error("No workbook created yet")
            return
        
        summary_sheet = self.workbook.create_sheet(sheet_name, 0)
        summary_sheet['A1'] = "API Discovery Summary"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        
        return summary_sheet
